import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from IPython.display import display

# Запрос даты у пользователя
date_str = input("Введите дату в формате DD.MM.YYYY: ")
date_obj = datetime.strptime(date_str, "%d.%m.%Y")
short_date = date_obj.strftime("%d.%m.%Y")

# Чтение файлов Excel
df = pd.read_excel(f'Трафик {short_date}.xlsx')
df_mag = pd.read_excel('Данные для расчета трафика.xlsx',
                       sheet_name='Магазины')

# ИЗМЕНЕНИЕ 1: Оптимизация обработки данных через цепочку методов
# БЫЛО: 10 отдельных операций с датафреймом
# СТАЛО: одна цепочка методов - код более читаемый и эффективный
df = (df.iloc[14:]
      .dropna(axis=1, how='all')
      .reset_index(drop=True)
      .drop([1, 2, 3, 8, 11, 12])
      .T
      # Замена df.columns = df.iloc[0]
      .pipe(lambda x: x.set_axis(x.iloc[0], axis=1))
      .iloc[1:]                                       # Замена df.drop([0])
      .reset_index(drop=True)
      .rename(columns={'Тип показателя': "Магазины"}))

# Удаляем лишние/закрытые магазины
StopList = ['РЦ "Победа" Ритейл',
            'РЦ «Родина» ИНЕТ',
            'РЦ «Родина» РИТЕЙЛ',
            '0005 Казань, Победа, ул. Закиева д.1',
            '0006 Нижнекамск, Якорь ул. Корабельная 44',
            '0008 Уфа, МИР просп. Октября, д. 4. кор. 1',
            '0010 Октябрьский, Верба, Ленина, д. 59/1',
            '0022 Ульяновск, ТЦ Самолет, просп.Ульяновский, д.1',
            '0023 Белебей, ТЦ Эссен, ул. им. Морозова, д.9',
            '0025 Ульяновск, ТЦ Пушкаревское, Московское ш,91',
            '0029 Челябинск, ТРК Горки, Артиллерийская, 136',
            '0033 Казань, ТРК Мега, пр-т. Победы, 141',
            '0035 Екатеринбург, ТЦ Радуга Парк, Репина, 94',
            '0044 Ижевск, ТЦ Аврора Парк, Удмуртская, 304',
            '0045 Ульяновск, КОРНЕР Аквамолл, Московское ш. 108',
            '0047 Москва, ТРЦ Павелецкая Плаза, Павелецкая, д.3',
            '0053 Казань, ТЦ Парк Хаус, Пр. Ямашева 46/33',
            '0054 Уфа, ТРЦ Планета, Энтузиастов, д. 20',
            '0072 Н.Новгород, ТРЦ Океанис, пр. Гагарина, 35',
            '0073 Москва, ТРЦ Щелковский, Щелковское шоссе, 75',
            '0080 Ярославль, ТРЦ Аура, Победы, 41',
            '0081 Пермь, ТРЦ Эспланада, Петропавловская, 73а',
            '0082 МО Химки, КОРНЕР ТЦ Лига, Ленинградское ш, 5',
            '0083 Москва, КОРНЕР МТК ЕвроПарк, Рублевское ш, 62',
            '0089 Новосибирск, КОРНЕР Аура, Военная, 5',
            '0102 Казань, КОРНЕР МЕГА, пр-т. Победы, 141',
            '0103 Уфа, КОРНЕР МЕГА, Рубежная, 174',
            '0104 Омск, КОРНЕР МЕГА, Бульвар Архитекторов, 35',
            '0105 Ростов-На Дону, КОРНЕР МЕГА, Аксайский, 23',
            '0107 МО Химки, КОРНЕР МЕГА, микрорайон ИКЕА, 2',
            '0106 Н. Новгород, КОРНЕР МЕГА, а/д Волга, Федяково',
            '0108 Москва, КОРНЕР МЕГА ТС, Калужское шоссе 21км',
            '0114 Новокузнецк, КОРНЕР ТРЦ Планета, ул.Доз, 10а']

# Заменяем пустые значения на 0
df = df.fillna(0)

# Удаляем магазины из стоп-листа
indexMag = df[(df['Магазины'].isin(StopList))].index
df.drop(indexMag, inplace=True)

# Генерируем возрастающие значения от 1 до количества строк
number = range(1, len(df) + 1)

# Добавляем новый столбец '№ п/п' в начало датафрейма
df.insert(0, '№ п/п', list(number))

# Проверяем корректность выгрузки данных в 1С по магазинам.
list_mag = df['Магазины'].values.tolist()
df_mag['Result'] = np.where(
    df_mag['Магазины'].isin(list_mag), 'Данные есть', 'Н/Д')
# Исправлено: создаем новый DataFrame
df_mag_check = df_mag[df_mag['Result'] == 'Н/Д']

# Сохраняем отчеты в Excel
df.to_excel(f"Отчет за день за_{short_date}.xlsx",
            sheet_name=f'Отчет за день за_{short_date}', index=False)
# Исправлено: используем правильный DataFrame
df_mag_check.to_excel(f"Отчет по магазинам за_{short_date}.xlsx", index=False)

# Функция на вычисление суммы по столбцу


def sum_col(ws, col, col_top=2, col_start=1, tight=False):
    col_len = ws.max_row
    if tight:
        # Находим последнюю непустую строку
        for row in range(ws.max_row, 0, -1):
            if ws[f'{col}{row}'].value is not None:
                col_len = row
                break
    ws[f'{col}{col_start}'] = f'=SUM({col}{col_top}:{col}{col_len})'

# Функция на вычисление среднего по столбцу


def avg_col(ws, col, col_top=2, col_start=1, tight=False):
    col_len = ws.max_row
    if tight:
        # Находим последнюю непустую строку
        for row in range(ws.max_row, 0, -1):
            if ws[f'{col}{row}'].value is not None:
                col_len = row
                break
    ws[f'{col}{col_start}'] = f'=AVERAGE({col}{col_top}:{col}{col_len})'

# ИЗМЕНЕНИЕ 2: Улучшенная функция на вычисление среднего по столбцу Е
# БЫЛО: Два отдельных цикла с ручным парсингом значений
# СТАЛО: Используем вспомогательную функцию для подсчета сумм - код проще и надежнее


def avg_col_f(ws, col1, col_top1=2, col2='D', col_top2=2):
    def sum_column(col, start_row):
        """Вспомогательная функция для подсчета суммы столбца"""
        total = 0
        for row in range(start_row, ws.max_row + 1):
            try:
                total += float(ws[f'{col}{row}'].value or 0)
            except (ValueError, TypeError):
                continue
        return total

    sum1 = sum_column(col1, col_top1)
    sum2 = sum_column(col2, col_top2)

    ws['F1'] = sum1 / sum2 if sum2 != 0 else 0

# ИЗМЕНЕНИЕ 3: Новая функция для применения форматирования
# БЫЛО: Множество повторяющихся циклов для каждого типа форматирования
# СТАЛО: Одна универсальная функция с конфигурируемыми правилами


def apply_formatting(ws, format_rules):
    """Применяет форматирование к листу Excel на основе переданных правил"""
    for rule in format_rules:
        cells = ws[rule['range']]
        for row in cells:
            for cell in row:
                for attr, value in rule['styles'].items():
                    setattr(cell, attr, value)


# Загружаем нужный файл эксель в работу
wb = load_workbook(f"Отчет за день за_{short_date}.xlsx")

# Загружаем нужный лист эксель в работу
ws = wb[f'Отчет за день за_{short_date}']

# Добавляем пустую строку сверху
ws.insert_rows(1)

# Проводим нужные вычисления
sum_col(ws, 'C', col_top=3, col_start=1)
sum_col(ws, 'D', col_top=3, col_start=1, tight=True)
avg_col(ws, 'E', col_top=3, col_start=1, tight=True)
avg_col_f(ws, 'C', col_top1=3, col2='D', col_top2=3)
sum_col(ws, 'G', col_top=3, col_start=1, tight=True)
sum_col(ws, 'H', col_top=3, col_start=1, tight=True)

# ИЗМЕНЕНИЕ 3 (продолжение): Конфигурация правил форматирования
# Все настройки форматирования в одном месте - легко изменять и поддерживать
format_rules = [
    # Жирный шрифт для заголовков (первые две строки)
    {
        'range': 'A1:H2',
        'styles': {'font': Font(bold=True, size=11)}
    },
    # Цвет фона для первой строки (основной заголовок)
    {
        'range': 'A1:H1',
        'styles': {'fill': PatternFill(start_color="FFE4C4", end_color="FFE4C4", fill_type="solid")}
    },
    # Цвет фона для второй строки (подзаголовок с фильтрами)
    {
        'range': 'A2:H2',
        'styles': {'fill': PatternFill(start_color="BCEE68", end_color="BCEE68", fill_type="solid")}
    },
    # Границы для всей таблицы
    {
        'range': f'A1:H{ws.max_row}',
        'styles': {'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))}
    }
]

# Применяем все правила форматирования одной функцией
apply_formatting(ws, format_rules)

# Автоматическая установка ширины столбцов
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if cell.value is not None and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    # Ограничиваем максимальную ширину
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Форматируем формат чисел в 1-й строчке
ws['C1'].number_format = '# ##0'
ws['D1'].number_format = '# ##0'
ws['E1'].number_format = '# ##0.00'
ws['F1'].number_format = '0.000'
ws['G1'].number_format = '# ##0'
ws['H1'].number_format = '# ##0'

# Определяем колонки для форматирования
columns_to_format = ['A', 'C', 'D', 'E', 'F', 'G', 'H']

# Проходим по заданным ячейкам
for col in columns_to_format:
    for row in range(1, ws.max_row + 1):
        cell = ws[f'{col}{row}']
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Устанавливаем фильтр на первую строку
ws.auto_filter.ref = f"A2:H{ws.max_row}"

# Сохраняем результаты изменений в файле эксель
wb.save(f'Отчет за день за_{short_date}.xlsx')

display(df)
df.info()
